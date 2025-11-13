"""Check what master data is in Excel vs existing CSV."""
import csv
from pathlib import Path
from openpyxl import load_workbook

# Load Excel
wb = load_workbook('Khavda Phase-3 Solar_Projects and SO Data (1).xlsx', read_only=True, data_only=True)
sheet = wb['Khavda-PhIII-Solar-SO Mapping']

# Read rows (skip 3 header rows)
rows = []
for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=1):
    if not any(row):
        break
    rows.append(row)

print(f"Excel has {len(rows)} data rows\n")

# Column mappings from introspection
# 0: Sr. No.
# 1: Cluster
# 2: Location
# 4: SPV
# 7: Project
# 15: Package
# 17: Vendor Code
# 18: Vendor Name
# 13: Project Definition Code > SPV

# Extract unique values
clusters = set()
locations = set()
spvs = set()
projects = set()
packages = set()
vendor_codes = set()
proj_defs = set()

for row in rows:
    if row[1]: clusters.add(str(row[1]).strip())
    if row[2]: locations.add(str(row[2]).strip())
    if row[4]: spvs.add(str(row[4]).strip())
    if row[7]: projects.add(str(row[7]).strip())
    if row[15]: packages.add(str(row[15]).strip())
    if row[17]: vendor_codes.add(str(row[17]).strip())
    if row[13]: proj_defs.add(str(row[13]).strip())

print("=== CLUSTERS ===")
print(f"Unique in Excel: {len(clusters)}")
for c in sorted(clusters):
    print(f"  - {c}")

print("\n=== LOCATIONS ===")
print(f"Unique in Excel: {len(locations)}")
for loc in sorted(locations):
    print(f"  - {loc}")

print("\n=== SPVS ===")
print(f"Unique in Excel: {len(spvs)}")
for s in sorted(spvs):
    print(f"  - {s}")

print("\n=== PROJECTS ===")
print(f"Unique in Excel: {len(projects)}")
for p in sorted(projects):
    print(f"  - {p}")

print("\n=== PACKAGES ===")
print(f"Unique in Excel: {len(packages)}")
for pkg in sorted(packages):
    print(f"  - {pkg}")

print("\n=== PROJECTDEFINITIONS (CODE) ===")
print(f"Unique in Excel: {len(proj_defs)}")
for pd in sorted(proj_defs)[:10]:
    print(f"  - {pd}")
if len(proj_defs) > 10:
    print(f"  ... and {len(proj_defs)-10} more")

print("\n=== VENDORS (CODE) ===")
print(f"Unique in Excel: {len(vendor_codes)}")
for vc in sorted(vendor_codes)[:15]:
    print(f"  - {vc}")
if len(vendor_codes) > 15:
    print(f"  ... and {len(vendor_codes)-15} more")

# Now check against existing
print("\n" + "="*60)
print("COMPARISON WITH EXISTING CSV")
print("="*60)

# Load existing
data_dir = Path('data')

def check_exists(entity, col, excel_vals):
    csv_file = data_dir / f'CCTECH.DRS.ENTITIES-{entity}.csv'
    if not csv_file.exists():
        return set(), set()
    
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        existing = {str(row.get(col, '')).strip().upper() for row in reader if row.get(col)}
    
    excel_upper = {v.upper() for v in excel_vals}
    already_exist = excel_upper & existing
    new_items = excel_upper - existing
    
    return already_exist, new_items

mappings = {
    'CLUSTERS': ('NAME', clusters),
    'LOCATIONS': ('NAME', locations),
    'SPVS': ('NAME', spvs),
    'PROJECTS': ('NAME', projects),
    'PACKAGES': ('NAME', packages),
    'VENDORS': ('CODE', vendor_codes),
    'PROJECTDEFINITIONS': ('CODE', proj_defs),
}

for entity, (col, excel_vals) in mappings.items():
    exists, new = check_exists(entity, col, excel_vals)
    print(f"\n{entity}:")
    print(f"  Already exist in CSV: {len(exists)}")
    if exists:
        for v in sorted(exists)[:5]:
            print(f"    ✓ {v}")
    print(f"  NEW (not in CSV): {len(new)}")
    if new:
        for v in sorted(new)[:5]:
            print(f"    + {v}")
        if len(new) > 5:
            print(f"    ... and {len(new)-5} more")
