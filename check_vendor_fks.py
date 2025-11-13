"""Analyze remaining VENDOR FK errors."""
import csv
from openpyxl import load_workbook
from pathlib import Path

# Load Excel
wb = load_workbook('Khavda Phase-3 Solar_Projects and SO Data (1).xlsx', read_only=True, data_only=True)
sheet = wb['Khavda-PhIII-Solar-SO Mapping']

# Read rows
rows = []
for row in sheet.iter_rows(min_row=4, values_only=True):
    if not any(row):
        break
    rows.append(row)

# Load existing + staged vendors
existing_vendors = set()
with open('data/CCTECH.DRS.ENTITIES-VENDORS.csv') as f:
    reader = csv.DictReader(f)
    for row in reader:
        existing_vendors.add(str(row['CODE']).strip())

staged_vendors = set()
if Path('staging/VENDORS.csv').exists():
    with open('staging/VENDORS.csv') as f:
        reader = csv.DictReader(f)
        for row in reader:
            staged_vendors.add(str(row['CODE']).strip())

all_vendors = existing_vendors | staged_vendors

# Check error rows 91, 92, 93
error_rows = [91, 92, 93, 94, 95, 100, 110, 120]
print("SERVICEORDERS VENDOR FK errors:\n")
missing_vendors = set()

for row_num in error_rows:
    row_idx = row_num - 1
    if row_idx < len(rows):
        row = rows[row_idx]
        vendor_code = str(row[17]).strip() if row[17] else ''
        so_num = str(row[16]).strip() if row[16] else ''
        print(f"  Row {row_num}: SO={so_num}, Vendor={vendor_code}")
        if vendor_code and vendor_code not in all_vendors:
            print(f"    ❌ Vendor '{vendor_code}' NOT AVAILABLE")
            missing_vendors.add(vendor_code)
        elif vendor_code:
            print(f"    ✓ Vendor '{vendor_code}' available")

if missing_vendors:
    print(f"\n⚠️  {len(missing_vendors)} missing vendor codes:")
    for v in sorted(missing_vendors):
        print(f"  - {v}")
else:
    print("\n✅ All vendors are available - FK errors must be from unique constraint violations")
