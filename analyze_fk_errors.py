"""Analyze FK resolution failures."""
import csv
import json
from pathlib import Path
from openpyxl import load_workbook

# Load schema
with open('build/schema.json') as f:
    schema = json.load(f)

# Load Excel
wb = load_workbook('Khavda Phase-3 Solar_Projects and SO Data (1).xlsx', read_only=True, data_only=True)
sheet = wb['Khavda-PhIII-Solar-SO Mapping']

# Read rows
rows = []
for row in sheet.iter_rows(min_row=4, values_only=True):
    if not any(row):
        break
    rows.append(row)

# Column 7 is Project
projects_in_excel = {str(row[7]).strip().upper() for row in rows if row[7]}

# Load existing PROJECTS
existing_projects = set()
with open('data/CCTECH.DRS.ENTITIES-PROJECTS.csv') as f:
    reader = csv.DictReader(f)
    for row in reader:
        existing_projects.add(row['NAME'].strip().upper())

# Load staged PROJECTS
staged_projects = set()
if Path('staging/PROJECTS.csv').exists():
    with open('staging/PROJECTS.csv') as f:
        reader = csv.DictReader(f)
        for row in reader:
            staged_projects.add(row['NAME'].strip().upper())

all_available = existing_projects | staged_projects

print("="*60)
print("PROJECTS ANALYSIS")
print("="*60)
print(f"\nIn Excel: {len(projects_in_excel)}")
print(f"Existing in CSV: {len(existing_projects)}")
print(f"Staged (NEW): {len(staged_projects)}")
print(f"Total available: {len(all_available)}")

missing = projects_in_excel - all_available
if missing:
    print(f"\n⚠️  MISSING (in Excel but not available): {len(missing)}")
    for p in sorted(missing):
        print(f"  - {p}")
else:
    print("\n✅ All projects from Excel are available!")

# Check specific errors from staging output
print("\n" + "="*60)
print("CHECKING ERROR ROWS")
print("="*60)

# The errors were at rows 14, 47, 50, 56, 61 for PROJECTDEFINITIONS
error_rows = [14, 47, 50, 56, 61]
print("\nPROJECTDEFINITIONS FK errors:")
for row_num in error_rows:
    row_idx = row_num - 1
    if row_idx < len(rows):
        row = rows[row_idx]
        project = str(row[7]).strip() if row[7] else ''
        proj_def = str(row[13]).strip() if row[13] else ''
        print(f"  Row {row_num}: Project='{project}', ProjDef='{proj_def}'")
        if project.upper() not in all_available:
            print(f"    ❌ Project '{project}' NOT AVAILABLE")

# Check SERVICEORDERS errors (rows 12, 13, 14, 15, 16)
print("\nSERVICEORDERS errors (first 5):")
error_rows_so = [12, 13, 14, 15, 16]
for row_num in error_rows_so:
    row_idx = row_num - 1
    if row_idx < len(rows):
        row = rows[row_idx]
        so_num = str(row[16]).strip() if row[16] else ''
        project = str(row[7]).strip() if row[7] else ''
        vendor = str(row[17]).strip() if row[17] else ''
        print(f"  Row {row_num}: SO='{so_num}', Project='{project}', Vendor='{vendor}'")
