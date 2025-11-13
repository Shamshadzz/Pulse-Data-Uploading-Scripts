"""
Introspect Excel workbook and return sheet names with headers.
"""
import os
from pathlib import Path
from typing import Dict, List, Any


def introspect_workbook(workbook_path: str, header_rows: int = 3) -> Dict[str, Any]:
    """
    Read Excel workbook and return sheet names with their headers.
    Handles multi-row hierarchical headers by merging them.
    
    Args:
        workbook_path: Path to Excel file
        header_rows: Number of header rows to process (default 3)
    
    Returns:
        {
            'workbook': filename,
            'sheets': [
                {
                    'name': 'SheetName',
                    'headers': ['Col1', 'Parent > Child', ...],
                    'raw_headers': [[row1], [row2], [row3]]
                },
                ...
            ]
        }
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install: pip install openpyxl")
    
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    
    result = {
        'workbook': os.path.basename(workbook_path),
        'sheets': []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Read multiple header rows
        raw_headers = []
        for row_idx in range(1, header_rows + 1):
            row_data = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
            if row_data:
                row_values = [str(cell).strip() if cell is not None else '' for cell in row_data]
                raw_headers.append(row_values)
        
        if not raw_headers:
            continue
        
        # Determine max columns
        max_cols = max(len(row) for row in raw_headers)
        
        # Normalize all rows to same length
        for row in raw_headers:
            while len(row) < max_cols:
                row.append('')
        
        # Merge hierarchical headers
        headers = []
        for col_idx in range(max_cols):
            parts = []
            for row_idx in range(len(raw_headers)):
                val = raw_headers[row_idx][col_idx].strip()
                if val:
                    parts.append(val)
            
            # Create hierarchical name
            if parts:
                # Use last part as primary, but keep hierarchy for disambiguation
                if len(parts) == 1:
                    header_name = parts[0]
                else:
                    # Parent > Child format
                    header_name = ' > '.join(parts)
            else:
                header_name = ''
            
            headers.append(header_name)
        
        # Remove empty trailing columns
        while headers and not headers[-1]:
            headers.pop()
        
        result['sheets'].append({
            'name': sheet_name,
            'headers': headers,
            'raw_headers': raw_headers
        })
    
    wb.close()
    return result


if __name__ == '__main__':
    import sys
    root = Path(__file__).parent.parent
    workbook = root / "Khavda Phase-3 Solar_Projects and SO Data (1).xlsx"
    
    if not workbook.exists():
        print(f"ERROR: Workbook not found: {workbook}", file=sys.stderr)
        sys.exit(1)
    
    info = introspect_workbook(str(workbook), header_rows=3)
    
    print(f"Workbook: {info['workbook']}")
    print(f"Found {len(info['sheets'])} sheets:\n")
    
    for sheet in info['sheets']:
        print(f"Sheet: {sheet['name']}")
        print(f"  Merged Headers ({len(sheet['headers'])}):")
        for idx, header in enumerate(sheet['headers'], 1):
            if header:
                print(f"    [{idx:2d}] {header}")
        print()
