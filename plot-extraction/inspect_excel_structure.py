"""
Inspect Excel file structure to understand the data layout.
This script examines the first Excel file to identify:
- Sheet names
- Header row count
- Column names and positions
- Sample data for MMS Table Names and Inverter Names
"""

import openpyxl
from pathlib import Path

def inspect_excel_file(file_path):
    """Inspect Excel file structure and print detailed information."""
    print(f"\n{'='*80}")
    print(f"Inspecting Excel File: {Path(file_path).name}")
    print(f"{'='*80}\n")
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # Print sheet names
    print(f"Sheet Names ({len(wb.sheetnames)} sheets):")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {sheet_name}")
    print()
    
    # Examine each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'-'*80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'-'*80}\n")
        
        ws = wb[sheet_name]
        
        # Get sheet dimensions
        print(f"Sheet Dimensions:")
        print(f"  Max Row: {ws.max_row}")
        print(f"  Max Column: {ws.max_column}")
        print()
        
        # Read first 10 rows to identify headers and data structure
        print("First 10 Rows:")
        print("-" * 120)
        
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):  # Limit to first 20 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    # Truncate long values
                    value_str = str(value)
                    if len(value_str) > 30:
                        value_str = value_str[:27] + "..."
                    row_data.append(f"Col{col_idx:02d}: {value_str}")
            
            if row_data:  # Only print non-empty rows
                print(f"Row {row_idx:02d}: {' | '.join(row_data)}")
        
        print()
        
        # Try to identify header rows
        print("Attempting to identify header structure:")
        potential_headers = []
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_values = [ws.cell(row=row_idx, column=col_idx).value 
                          for col_idx in range(1, ws.max_column + 1)]
            non_empty = [v for v in row_values if v is not None and str(v).strip()]
            if non_empty:
                potential_headers.append((row_idx, row_values))
                print(f"  Row {row_idx}: {len(non_empty)} non-empty cells")
        print()
        
        # Look for columns containing specific patterns
        print("Searching for MMS Table Names and Inverter Names columns:")
        print("-" * 120)
        
        # Sample first 20 data rows after potential headers
        start_row = len(potential_headers) + 1
        sample_data = {}
        
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # Get header (assuming it's in one of the first rows)
            header_values = []
            for row_idx in range(1, min(4, ws.max_row + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    header_values.append(str(val).strip())
            
            header = " | ".join(header_values) if header_values else f"Col{col_idx}"
            
            # Get sample data
            data_samples = []
            for row_idx in range(start_row, min(start_row + 20, ws.max_row + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and str(val).strip():
                    data_samples.append(str(val).strip())
            
            # Check if this column contains table/inverter patterns
            has_table_pattern = any('R' in s and 'T' in s for s in data_samples[:10])
            has_inverter_pattern = any(s.startswith('I') or '-I' in s for s in data_samples[:10])
            has_block_pattern = any('B' in s and '-' in s for s in data_samples[:10])
            
            if has_table_pattern or has_inverter_pattern or has_block_pattern:
                print(f"\n📍 Column {col_letter} (Index {col_idx}):")
                print(f"   Header: {header[:80]}")
                print(f"   Sample Data (first 10 non-empty):")
                for i, sample in enumerate(data_samples[:10], 1):
                    # Check pattern
                    pattern_type = ""
                    if 'R' in sample and ('T' in sample or 'S' in sample):
                        pattern_type = " [TABLE]"
                    elif sample.startswith('I') or '-I' in sample:
                        pattern_type = " [INVERTER]"
                    print(f"      {i:2d}. {sample}{pattern_type}")
        
        print("\n" + "="*80)
    
    wb.close()
    print("\n✅ Inspection complete!\n")

if __name__ == "__main__":
    # Path to the Excel file
    excel_file = r"c:\Users\Shamshad choudhary\Documents\Pulse-Data-Uploading-Scripts\plot-extraction\drawing_data\A16a - 50 MW\603C-LT Cable Routing-A16a-BL01-R0-30032025_DWGData.xlsx"
    
    try:
        inspect_excel_file(excel_file)
    except Exception as e:
        print(f"❌ Error inspecting file: {e}")
        import traceback
        traceback.print_exc()
