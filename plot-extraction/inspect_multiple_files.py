"""
Inspect multiple Excel files to confirm consistent structure across blocks.
"""

import openpyxl
from pathlib import Path

def inspect_file_quick(file_path):
    """Quick inspection of Excel file structure."""
    print(f"\n{'='*80}")
    print(f"File: {Path(file_path).name}")
    print(f"{'='*80}")
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Sheet: {sheet_name}")
            print(f"  Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Check header
            header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            print(f"  Headers: {header_row}")
            
            # Sample first few data rows
            print(f"  Sample Data (rows 2-5):")
            for row_idx in range(2, min(6, ws.max_row + 1)):
                row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
                print(f"    Row {row_idx}: {row_data}")
        
        wb.close()
        
    except Exception as e:
        print(f"  ❌ Error: {e}")

if __name__ == "__main__":
    base_path = Path(r"c:\Users\Shamshad choudhary\Documents\Pulse-Data-Uploading-Scripts\plot-extraction\drawing_data\A16a - 50 MW")
    
    excel_files = list(base_path.glob("*.xlsx"))
    
    print(f"Found {len(excel_files)} Excel files in {base_path.name}")
    
    for file_path in excel_files:
        inspect_file_quick(file_path)
